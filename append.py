import pandas as pd
from openpyxl import load_workbook

# New data to append
new_data = pd.DataFrame({
    'Name': ['Alice', 'Bob'],
    'Score': [88, 92]
})

# File and target sheet
file_path = 'eg.xlsx'
sheet_name = 'Sheet2'

# Load workbook to find last row
book = load_workbook(file_path)
last_row = book[sheet_name].max_row

# Append using pandas ExcelWriter without setting writer.book
with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    new_data.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=last_row)
